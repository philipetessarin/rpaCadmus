import time
import openpyxl
from selenium.webdriver import Keys

from enviar_email import EnviarEmail

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options

class BuscaVagas:
    def __init__(self):
        self.chrome_opcoes = Options()
        self.argumentos = ['--headless', '--no-sandbox', '--disable-extensions']

        for argumento in self.argumentos:
            self.chrome_opcoes.add_argument(argumento)

        self.driver = webdriver.Chrome('./chromedriver.exe', options=self.chrome_opcoes)

    def iniciar(self):
        self.driver.get("https://cadmus.com.br/vagas-tecnologia/")
        time.sleep(4)

        self.criar_planilha()
        self.encontrar_vagas_na_pagina()

    def criar_planilha(self) -> None:
        self.book = openpyxl.Workbook()

        self.planilha_vagas = self.book['Sheet']
        self.planilha_vagas.cell(row=1, column=1, value='Nome')
        self.planilha_vagas.cell(row=1, column=2, value='Local')
        self.planilha_vagas.cell(row=1, column=3, value='Descrição')

    def encontrar_vagas_na_pagina(self) -> None:
        try:
            self.qtd_vagas = len(self.driver.find_elements(by=By.XPATH, value='//h3'))

            self.descricoes = []

            for indice in range(0, self.qtd_vagas):
                print(f'Capturando Descrição {indice + 1} de {self.qtd_vagas}...')
                self.descricao = self.driver.find_element(by=By.XPATH,
                                                          value=f'//a[@onclick="javascript:AbrirOportunidade(\'{indice}\');"]')
                self.descricao.send_keys(Keys.ENTER)
                time.sleep(1)
                self.descricao = self.driver.find_element(by=By.XPATH, value='//*[@id="boxVaga"]/p').get_attribute(
                    "innerText").replace("\n", " ")
                self.descricoes.append(self.descricao)
                # time.sleep(2)
                self.driver.get("https://cadmus.com.br/vagas-tecnologia/")

            self.titulos = self.driver.find_elements(by=By.XPATH, value='//h3')
            self.localizacoes = self.driver.find_elements(by=By.XPATH, value='//p[@class="local"]')

            self.armazenar_vagas_na_planilha()
        except Exception as erro:
            print('Erro ao extrair informações.')

    def armazenar_vagas_na_planilha(self) -> None:
        print('Adicionando informações das vagas na planilha...')
        for titulo, local, descricao in zip(self.titulos, self.localizacoes, self.descricoes):
            nova_linha = [titulo.get_attribute("innerText"), local.get_attribute("innerText"), descricao]
            self.planilha_vagas.append(nova_linha)

        self.driver.close()
        self.book.save('Vagas_Cadmus.xlsx')

        print('Enviando email...')
        email = EnviarEmail()
        email.envio()
        print('Processo finalizado com sucesso.')

if __name__ == '__main__':
    encontrar_vagas = BuscaVagas()
    encontrar_vagas.iniciar()
